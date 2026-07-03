"""XLSX conversion handler."""


class XlsxConverter:
    """Convert XLSX to Markdown, HTML, CSV, JSON, and text."""

    def convert(self, input_path, target_format, output_path):
        if target_format == "md":
            self._to_markdown(input_path, output_path)
        elif target_format == "html":
            self._to_html(input_path, output_path)
        elif target_format == "csv":
            self._to_csv(input_path, output_path)
        elif target_format == "json":
            self._to_json(input_path, output_path)
        elif target_format == "txt":
            self._to_text(input_path, output_path)
        else:
            raise ValueError(f"Unsupported XLSX target format: {target_format}")

    def _read_sheet(self, input_path):
        from openpyxl import load_workbook
        wb = load_workbook(input_path, data_only=True)
        return wb.active

    def _rows(self, ws):
        for row in ws.iter_rows(values_only=True):
            yield [str(cell) if cell is not None else "" for cell in row]

    def _to_markdown(self, input_path, output_path):
        ws = self._read_sheet(input_path)
        rows = list(self._rows(ws))
        if not rows:
            with open(output_path, "w", encoding="utf-8") as f:
                f.write("")
            return

        lines = []
        lines.append("| " + " | ".join(rows[0]) + " |")
        lines.append("| " + " | ".join(["---"] * len(rows[0])) + " |")
        for row in rows[1:]:
            while len(row) < len(rows[0]):
                row.append("")
            lines.append("| " + " | ".join(row[:len(rows[0])]) + " |")

        with open(output_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))

    def _to_html(self, input_path, output_path):
        ws = self._read_sheet(input_path)
        rows = list(self._rows(ws))
        html = ["<!DOCTYPE html>", "<html><head><meta charset=\"utf-8\"></head><body>", "<table border=\"1\">"]
        for row in rows:
            html.append("<tr>")
            for cell in row:
                escaped = cell.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                html.append(f"<td>{escaped}</td>")
            html.append("</tr>")
        html.extend(["</table>", "</body></html>"])
        with open(output_path, "w", encoding="utf-8") as f:
            f.write("\n".join(html))

    def _to_csv(self, input_path, output_path):
        import csv
        ws = self._read_sheet(input_path)
        with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow([cell if cell is not None else "" for cell in row])

    def _to_json(self, input_path, output_path):
        import json
        ws = self._read_sheet(input_path)
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            data = []
        else:
            headers = [str(h) if h is not None else "" for h in rows[0]]
            data = []
            for row in rows[1:]:
                data.append({
                    headers[i]: (str(row[i]) if i < len(row) and row[i] is not None else "")
                    for i in range(len(headers))
                })
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def _to_text(self, input_path, output_path):
        ws = self._read_sheet(input_path)
        lines = ["\t".join(row) for row in self._rows(ws)]
        with open(output_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
